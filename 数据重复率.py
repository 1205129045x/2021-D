import pandas as pd
feature_train=pd.read_csv('Molecular_Descriptor.csv',index_col='SMILES')
a=[]
for  s in feature_train.columns:
    a.append(feature_train[s].value_counts().iloc[0])
new_list = [x/1974 for x in a]
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter, column_index_from_string

# 向sheetobj中的columnname列从start_row开始写入listdata
def insert_listdata_to_column(sheetobj,listdata,column_name,start_row=2):
    colindex = column_index_from_string(column_name)
    for rowindex in range(start_row, start_row + len(data)):
        val = data[rowindex - start_row]
        try:
            sheet.cell(row = rowindex,column = colindex,value = val)
        except:
            val = ILLEGAL_CHARACTERS_RE.sub(r'',val)
            sheet.cell(row = rowindex,column = colindex,value = val)

column_mapping = {
	"a":range(1,730),
	"b":new_list,
    "c":range(1,730),
}


#创建新的表格
wb = openpyxl.Workbook()
sheet = wb.active
for key,val in column_mapping.items():
    col_name = key
    data = val
    insert_listdata_to_column(sheet,data,col_name,2)
sheet['a1']='序号'
sheet['b1']='重复率'
sheet['c1']='特征重要度'
wb.save("ttt.xlsx")


